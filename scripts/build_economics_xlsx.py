"""Собрать docs/biznes/ЭКОНОМИКА.xlsx из шаблона с формулами."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "biznes" / "ЭКОНОМИКА.xlsx"

BLUE = Font(name="Arial", color="0000FF")
BLACK = Font(name="Arial", color="000000")
BOLD = Font(name="Arial", bold=True)
HDR = Font(name="Arial", bold=True, size=11)
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
RUB = '#,##0" ₽"'
PCT = "0.0%"


def style_input(cell, value):
    cell.value = value
    cell.font = BLUE
    cell.fill = INPUT_FILL


def style_formula(cell, formula):
    cell.value = formula
    cell.font = BLACK


def set_width(ws, widths: dict[int, float]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def sheet_readme(wb: Workbook):
    ws = wb.active
    ws.title = "Как пользоваться"
    rows = [
        ("СоздайСвоюПесню — экономика", ""),
        ("", ""),
        ("Жёлтые ячейки", "Ваши цифры — меняйте сами"),
        ("Синий текст", "Тоже ввод (оценки до уточнения)"),
        ("Чёрные формулы", "Пересчитываются автоматически"),
        ("", ""),
        ("Листы:", ""),
        ("Ввод", "Главные допущения (курс, API, налоги)"),
        ("Фикс", "Постоянные расходы ₽/мес"),
        ("Пакеты", "Текущие тарифы и маржа"),
        ("Калькулятор", "Точка безубыточности"),
        ("Реклама", "CAC / LTV / пробные"),
        ("Балансы", "Журнал пополнений API"),
        ("Подписка", "Черновик планов"),
        ("", ""),
        ("Текст для ассистента:", "docs/biznes/ЭКОНОМИКА.txt"),
        ("Обновлено:", "10 июля 2026"),
    ]
    for r, (a, b) in enumerate(rows, 1):
        ws.cell(r, 1, a).font = BOLD if a and not b and r in (1, 7) else Font(name="Arial")
        ws.cell(r, 2, b)
    set_width(ws, {1: 28, 2: 42})


def sheet_vvod(wb: Workbook):
    ws = wb.create_sheet("Ввод")
    ws["A1"] = "Допущения (меняйте жёлтые/синие ячейки)"
    ws["A1"].font = HDR
    labels = [
        ("Курс USD → RUB", "usd_rub", 95),
        ("Suno + ApiPass, ₽/нота (оценка)", "suno_rub", 6),
        ("YandexGPT, ₽/нота", "yandex_rub", 5),
        ("S3, ₽/нота", "s3_rub", 0.5),
        ("Брак/retry, % к API", "waste_pct", 0.12),
        ("Эквайринг GetPlatinum, %", "gp_pct", 0.03),
        ("НПД, % с выручки", "npd_pct", 0.04),
        ("Платных нот в месяц (прогноз)", "notes_month", 50),
        ("Пробных генераций в месяц", "trials_month", 100),
        ("Себест. пробной, ₽", "trial_cost", 15),
    ]
    ws["A3"] = "Параметр"
    ws["B3"] = "Значение"
    ws["C3"] = "Ключ"
    for c in range(1, 4):
        ws.cell(3, c).font = BOLD
    for i, (label, key, default) in enumerate(labels, 4):
        ws.cell(i, 1, label)
        style_input(ws.cell(i, 2), default)
        ws.cell(i, 3, key)
        if "pct" in key:
            ws.cell(i, 2).number_format = PCT
        elif key == "usd_rub":
            ws.cell(i, 2).number_format = "#,##0.00"
        elif key in ("notes_month", "trials_month"):
            ws.cell(i, 2).number_format = "#,##0"
        else:
            ws.cell(i, 2).number_format = RUB
    ws["A15"] = "Переменная себест. 1 ноты, ₽"
    ws["A15"].font = BOLD
    style_formula(ws["B15"], "=B5+B6+B7+(B5+B6+B7)*B9")
    ws["B15"].number_format = RUB
    ws["A16"] = "Итого фикс ₽/мес (с листа Фикс)"
    ws["A16"].font = BOLD
    style_formula(ws["B16"], "=Фикс!B12")
    ws["B16"].number_format = RUB
    ws["A17"] = "Фикс на 1 ноту, ₽"
    style_formula(ws["B17"], "=IF(B11>0,B16/B11,0)")
    ws["B17"].number_format = RUB
    set_width(ws, {1: 36, 2: 14, 3: 12})


def sheet_fix(wb: Workbook):
    ws = wb.create_sheet("Фикс")
    ws["A1"] = "Постоянные расходы, ₽/мес"
    ws["A1"].font = HDR
    items = [
        ("VPS 195.19.20.245", 0),
        ("Домен sozdaipesnu.ru", 0),
        ("Почта REG.RU Mail-1", 0),
        ("S3 REG.RU", 0),
        ("Yandex Cloud (мин.)", 0),
        ("Прочее", 0),
    ]
    ws["A3"] = "Статья"
    ws["B3"] = "₽/мес"
    ws["C3"] = "Примечание"
    for c in range(1, 4):
        ws.cell(3, c).font = BOLD
    for i, (name, val) in enumerate(items, 4):
        ws.cell(i, 1, name)
        style_input(ws.cell(i, 2), val)
        ws.cell(i, 2).number_format = RUB
        ws.cell(i, 3, "заполнить")
    ws["A11"] = "CAPEX разово (не в месяц)"
    ws["A12"] = "ИТОГО фикс ₽/мес"
    ws["A12"].font = BOLD
    style_formula(ws["B12"], "=SUM(B4:B9)")
    ws["B12"].number_format = RUB
    set_width(ws, {1: 28, 2: 12, 3: 24})


def sheet_pakety(wb: Workbook):
    ws = wb.create_sheet("Пакеты")
    ws["A1"] = "Пакеты нот (как на сайте v2.11.9)"
    ws["A1"].font = HDR
    headers = ["ID", "Нот", "Цена ₽", "₽/нота", "Себест. ₽", "Маржа ₽", "Маржа %"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h).font = BOLD
    data = [
        ("notes_1", 1, 299),
        ("notes_3", 3, 749),
        ("notes_5", 5, 1199),
        ("notes_10", 10, 1799),
    ]
    for i, (pid, notes, price) in enumerate(data, 4):
        ws.cell(i, 1, pid)
        ws.cell(i, 2, notes)
        style_input(ws.cell(i, 3), price)
        ws.cell(i, 3).number_format = RUB
        style_formula(ws.cell(i, 4), f"=C{i}/B{i}")
        ws.cell(i, 4).number_format = RUB
        style_formula(
            ws.cell(i, 5),
            f"=B{i}*(Ввод!$B$15+Ввод!$B$17)+C{i}*(Ввод!$B$10+Ввод!$B$11)",
        )
        ws.cell(i, 5).number_format = RUB
        style_formula(ws.cell(i, 6), f"=C{i}-E{i}")
        ws.cell(i, 6).number_format = RUB
        style_formula(ws.cell(i, 7), f"=IF(C{i}>0,F{i}/C{i},0)")
        ws.cell(i, 7).number_format = PCT
    ws["A9"] = "Жёсткий пол ₽/нота (ориентир)"
    style_input(ws["B9"], 45)
    ws["B9"].number_format = RUB
    set_width(ws, {1: 12, 2: 8, 3: 12, 4: 12, 5: 12, 6: 12, 7: 10})


def sheet_kalk(wb: Workbook):
    ws = wb.create_sheet("Калькулятор")
    ws["A1"] = "Сводка месяца"
    ws["A1"].font = HDR
    rows = [
        ("Фикс ₽/мес", "=Фикс!B12"),
        ("Пробные × себест.", "=Ввод!B12*Ввод!B13"),
        ("Реклама ₽/мес", None, "input"),
        ("Итого затрат до маржи", "=B4+B5+B6"),
        ("Средняя цена ноты ₽", None, "input"),
        ("Переменная + фикс/нота ₽", "=Ввод!B15+Ввод!B17"),
        ("Маржа на ноту ₽", "=B9*(1-Ввод!B10-Ввод!B11)-B10"),
        ("Нот для безубыточности", "=IF(B11>0,B7/B11,0)"),
    ]
    ws["A3"] = "Показатель"
    ws["B3"] = "Значение"
    ws["A3"].font = ws["B3"].font = BOLD
    for i, (label, formula, *extra) in enumerate(rows, 4):
        ws.cell(i, 1, label)
        if extra and extra[0] == "input":
            style_input(ws.cell(i, 2), 10000 if "Реклама" in label else 220)
        else:
            style_formula(ws.cell(i, 2), formula)
        if i >= 7:
            ws.cell(i, 2).number_format = RUB if i != 11 else "#,##0"
    ws["B11"].number_format = "#,##0"
    set_width(ws, {1: 32, 2: 16})


def sheet_reklama(wb: Workbook):
    ws = wb.create_sheet("Реклама")
    ws["A1"] = "CAC / LTV"
    ws["A1"].font = HDR
    ws["A3"] = "Рекламный бюджет ₽"
    style_input(ws["B3"], 10000)
    ws["B3"].number_format = RUB
    ws["A4"] = "Платящих клиентов"
    style_input(ws["B4"], 20)
    ws["A5"] = "CAC ₽"
    style_formula(ws["B5"], "=IF(B4>0,B3/B4,0)")
    ws["B5"].number_format = RUB
    ws["A6"] = "Средний чек ₽"
    style_input(ws["B6"], 500)
    ws["B6"].number_format = RUB
    ws["A7"] = "Покупок на клиента"
    style_input(ws["B7"], 1.5)
    ws["A8"] = "LTV ₽"
    style_formula(ws["B8"], "=B6*B7")
    ws["B8"].number_format = RUB
    ws["A9"] = "LTV / CAC"
    style_formula(ws["B9"], "=IF(B5>0,B8/B5,0)")
    ws["B9"].number_format = "0.0"
    ws["A10"] = "Ок? (LTV > 3×CAC)"
    style_formula(ws["B10"], '=IF(B9>=3,"ДА","НЕТ")')
    set_width(ws, {1: 28, 2: 14})


def sheet_balansy(wb: Workbook):
    ws = wb.create_sheet("Балансы")
    ws["A1"] = "Журнал пополнений и алерты"
    ws["A1"].font = HDR
    headers = ["Дата", "Сервис", "Сумма", "Баланс после", "Алерт если", "Примечание"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h).font = BOLD
    services = [
        ("sunoapi.org", "credits < 100"),
        ("ApiPass", "$ < 5"),
        ("Yandex Cloud", "₽ < 100"),
        ("REG.RU S3", "квота 80%"),
    ]
    for i, (svc, alert) in enumerate(services, 4):
        ws.cell(i, 2, svc)
        ws.cell(i, 5, alert)
    for r in range(4, 14):
        for c in range(1, 7):
            if ws.cell(r, c).value is None and c not in (2, 5):
                ws.cell(r, c, "")
    set_width(ws, {1: 12, 2: 14, 3: 12, 4: 14, 5: 14, 6: 24})


def sheet_podpiska(wb: Workbook):
    ws = wb.create_sheet("Подписка")
    ws["A1"] = "Черновик подписки"
    ws["A1"].font = HDR
    headers = ["План", "Нот/мес", "Цена ₽", "₽/нота", "Сгорают?"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h).font = BOLD
    plans = [
        ("Старт", 5, 499, "да"),
        ("Базовый", 15, 1199, "да"),
        ("Про", 40, 2499, "да"),
    ]
    for i, (name, notes, price, burn) in enumerate(plans, 4):
        ws.cell(i, 1, name)
        ws.cell(i, 2, notes)
        style_input(ws.cell(i, 3), price)
        ws.cell(i, 3).number_format = RUB
        style_formula(ws.cell(i, 4), f"=C{i}/B{i}")
        ws.cell(i, 4).number_format = RUB
        ws.cell(i, 5, burn)
    set_width(ws, {1: 12, 2: 10, 3: 12, 4: 12, 5: 12})


def main():
    wb = Workbook()
    sheet_readme(wb)
    sheet_vvod(wb)
    sheet_fix(wb)
    sheet_pakety(wb)
    sheet_kalk(wb)
    sheet_reklama(wb)
    sheet_balansy(wb)
    sheet_podpiska(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"OK: {OUT}")


if __name__ == "__main__":
    main()